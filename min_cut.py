import openpyxl
import networkx as nx
from numpy import sqrt
import matplotlib.pyplot as plt
from networkx.algorithms.approximation import steinertree

# функция проверки числа направлений связи
def inter_num_direct():
    while True:
        num = input('Введите число: ')
        if num.isdecimal() and int(num) > 0:
            return \
                int(num)
            break
        else:
            print('Вы должны ввести целое число, больше нуля, пропробуте снова')

# функция проверки существования узла в графе
def inter_node():
    print('Перечень возможных Узлов %s' %list(G.nodes))
    node = str(input('Укажите узел: '))
    if node in G.nodes:
        return node
    else:
        print('Вы должны указать, номер Узла из перечняя возможных, пропробуте снова')

# открываю файл с координатами узлов (колонки обозначают: первая название узла, вторая координату по X,
# третья координата по Y)
wb = openpyxl.load_workbook(filename='point.xlsx')
# открываю активный лист
ws = wb.active
# создаю пустой список
data = []
# определяю количество узлов по количеству заполненных строк
N_point = ws.max_row
# собираю значения всех ячеек в список data
for row in ws.values:
   for value in row:
       data.append(value)

# создаю словарь для координат узлов
Nodes = {}
for i in range(0, len(data), 3):
    # заполняю словарь: первое значение ключ - номер узла (с нуля) втрое и третье заначение координаты
    Nodes.setdefault(str(data[i]), (data[i+1], data[i+2]))

# открываю файл с c матрицей инценденций (единица есть ребро, ноль нет ребра, по диагонале считать не нужно)
wb = openpyxl.load_workbook(filename='i-matrix.xlsx')
# открываю активный лист
ws = wb.active
# создаю пустой список
data = []
# проверяю соответствие перечня узлов матрице инценденций
if N_point != ws.max_row:
    print('матрица инценденций не соответствует количеству узлов')
    quit()
N_point = ws.max_row

# собираю значения всех ячеек в список data
for row in ws.values:
   for value in row:
       data.append(value)

# собираю список линий из данных матрицы инценденций
Edges = []
for i in range(0, N_point):
    l = i*5
    for j in range(0, N_point):
        if data[l+j] == 1:
            Edge = (str(i), str(j))
            Edges.append(Edge)

# Создаю граф
G = nx.Graph()
# Добавляю в граф сформированный список узлов
G.add_nodes_from(Nodes)
# Добавляю в граф сформированный список линий
G.add_edges_from(Edges)

# расчет длин линий
Lines = []
for edge in G.edges:
    distance = sqrt((Nodes[edge[0]][0] - Nodes[edge[1]][0]) ** 2 + (Nodes[edge[0]][1] - Nodes[edge[1]][1]) ** 2)
    # расчет длин ребер между узлами
    line = (edge[0], edge[1], distance)
    Lines.append(line)
G.add_weighted_edges_from(Lines)
# добавляем длины линий в граф

# определяю узлы для направлений связи
print('Укажите количетво направлений связи для расчета')
num_direct = inter_num_direct()
direct_node = []
for num in range(num_direct):
    print('Укажите первый узел корреспондирующей пары №%i' %num)
    direct_node.append(inter_node())
    print('Укажите втрой узел корреспондирующей пары №%i' %num)
    direct_node.append(inter_node())

# находим список линий составляющих минальное дерево Штейнера для корреспондирующих пар узлов
edge_list = list(nx.algorithms.approximation.steinertree.steiner_tree(G, direct_node).edges)

# расчет протяженности списка линий
summa_edge = 0
for edge in edge_list:
    summa_edge += float(G.edges[edge]['weight'])

# сохраняем результат в файл exel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'результат'
ws['A1'] = 'Протяженность линий'
ws['A2'] = summa_edge
ws['A3'] = 'Перечень линий'
res_edge_list = ''
I = 3
for i in range(0, len(edge_list)):
   I = I + 1
   ws['A'+str(I)] = str(edge_list[i])
wb.save(filename='Results.xlsx')

# выводим графически результат
plt.subplot(121)
plt.title('Исходные данные: \n Количество узлов - %i.' %N_point) # печатаю заголовок рисунка
nx.draw(G, pos=Nodes, with_labels=True) # рисую граф с названиями и позициями узлов и линий
plt.subplot(122)
plt.title('Результат расчета: \n Перечень линий - %s '
          '\n Протяженность линий - %f км' % (edge_list, summa_edge)) # печатаю заголовок рисунка
nx.draw(G, pos=Nodes, with_labels=True) # рисую граф с названиями и позициями узлов и линий
nx.draw_networkx_edges(G, pos=Nodes, edgelist=edge_list, width=3.0, edge_color='r')# рисую линии на графе
plt.draw() # формирую рисунок
plt.show() # вывожу рисунок


